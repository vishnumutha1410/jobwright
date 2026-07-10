"""Excel tracker + dedupe/history layer (the project's lightweight 'database').

Sheets:
  - Jobs:    one row per active, matched job with all fields.
  - History: every job_id ever seen (so we never notify/draft twice).
"""
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from config import TRACKER_PATH

JOB_HEADERS = [
    "job_id", "company", "title", "location", "salary", "employment_type",
    "experience", "req_skills", "match_score", "priority", "posted_days_ago",
    "why_match", "skill_gaps", "note", "interview_tips", "apply_url",
    "listing_url", "date_found", "status",
]
HEAD_FILL = PatternFill("solid", start_color="1F3864")
WHITE = Font(color="FFFFFF", bold=True)


def _ensure(path):
    if os.path.exists(path):
        return load_workbook(path)
    wb = Workbook()
    js = wb.active
    js.title = "Jobs"
    js.append(JOB_HEADERS)
    hs = wb.create_sheet("History")
    hs.append(["job_id", "company", "title", "apply_url", "first_seen", "drafted"])
    for sheet in (js, hs):
        for c in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=c)
            cell.fill = HEAD_FILL
            cell.font = WHITE
            cell.alignment = Alignment(horizontal="center")
    wb.save(path)
    return wb


def seen_ids(path=TRACKER_PATH) -> set:
    """Return job_ids already recorded in History (for dedupe)."""
    wb = _ensure(path)
    hs = wb["History"]
    return {row[0].value for row in hs.iter_rows(min_row=2) if row[0].value}


def append_jobs(jobs, path=TRACKER_PATH):
    """Append new jobs to Jobs + History and save. Returns count added."""
    wb = _ensure(path)
    js, hs = wb["Jobs"], wb["History"]
    from datetime import date
    added = 0
    for j in jobs:
        js.append(_row(j))
        hs.append([j.job_id, j.company, j.title, j.apply_url, date.today().isoformat(), "Yes"])
        added += 1
    _autosize(js)
    wb.save(path)
    return added


def _row(j):
    d = {
        "job_id": j.job_id, "company": j.company, "title": j.title, "location": j.location,
        "salary": j.salary, "employment_type": j.employment_type, "experience": j.experience,
        "req_skills": ", ".join(j.req_skills), "match_score": j.match_score, "priority": j.priority,
        "posted_days_ago": j.posted_days_ago, "why_match": j.why_match, "skill_gaps": j.skill_gaps,
        "note": j.note, "interview_tips": j.interview_tips, "apply_url": j.apply_url,
        "listing_url": j.listing_url, "date_found": j.date_found, "status": "Not applied",
    }
    return [d[h] for h in JOB_HEADERS]


def _autosize(sheet, cap=60):
    for col in sheet.columns:
        width = min(cap, max((len(str(c.value)) for c in col if c.value), default=10) + 2)
        sheet.column_dimensions[col[0].column_letter].width = width
